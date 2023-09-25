import re
import openpyxl as op
from tqdm import tqdm


# Importing Buyers and Sellers data for training
wb=op.load_workbook(r"NLP\training_intent_samples.xlsx")
ws=wb.active

buyers= []
buyer_verbs=[]
products=[]
sellers = []
seller_verbs=[]
ad_camp=[]
negations=[]
intent_suffixes=[]

# Extracting buyers and seller lists of words or phrases to predict the intents
for row in ws.iter_rows():
    if row[0].value is not None:
        buyers.append(str(row[0].value))
        
    if row[1].value is not None:
        buyer_verbs.append(str(row[1].value))

    if row[2].value is not None:
        products.append(str(row[2].value))
        
    if row[3].value is not None:
        sellers.append(str(row[3].value))
        
    if row[4].value is not None:
        seller_verbs.append(str(row[4].value))
        
    if row[5].value is not None:
        ad_camp.append(str(row[5].value))
        
    if row[6].value is not None:
        negations.append(str(row[6].value))
        
    if row[7].value is not None:
        intent_suffixes.append(str(row[7].value))
        

# Creating patterns for regular expression for buyer and seller intents
buyer_pattern = r'\b(?:' + '|'.join(map(re.escape, buyers)) + r')\b'
buyer_verbs_pattern = r'\b(?:' + '|'.join(map(re.escape, buyer_verbs)) + r')\b'
products_pattern = r'\b(?:' + '|'.join(map(re.escape, products)) + r')'
sellers_pattern = r'\b(?:' + '|'.join(map(re.escape, sellers)) + r')\b'
seller_verbs_pattern = r'\b(?:' + '|'.join(map(re.escape, seller_verbs)) + r')\b'
ad_camp_pattern = r'\b(?:' + '|'.join(map(re.escape, ad_camp)) + r')\b'
negations_pattern = r'\b(?:' + '|'.join(map(re.escape, negations)) + r')\b' 
intent_suffixes_pattern = r'\b(?:' + '|'.join(map(re.escape, intent_suffixes)) + r')' 

wb.close()
# -------------------------------------------------------

# Reading training Data
# Choice 
choice=int(input('''For sentence input select "1" 
For txt(s) files select "2" 
>>> ''').strip())

training_files_n_labels=[]


def evaluate(line,label):
    text=line.strip()
    
    # setting row back to 0
    row=0 
    wb2=op.load_workbook(r"NLP\dataset_trained.xlsx")
    ws2 = wb2.active
    # Adding Header to all column in trained dataset
    if ws2.cell(row+1,1).value!="buyer":
        ws2.cell(row+1,1).value="buyer"
        ws2.cell(row+1,2).value="seller"
        ws2.cell(row+1,3).value="seller_verb"
        ws2.cell(row+1,4).value="buyer_verb"
        ws2.cell(row+1,5).value="products"
        ws2.cell(row+1,6).value="ad_campaign"
        ws2.cell(row+1,7).value="intent_sufixes"
        ws2.cell(row+1,8).value="negations"
        ws2.cell(row+1,9).value="label"
              
    row= ws2.max_row # setting row to next empty cell
    
    # Finding all occurrences using the re module for patterns
    buyer_match = re.findall(buyer_pattern, text, flags=re.IGNORECASE)
    buyer_verbs_match = re.findall(buyer_verbs_pattern, text, flags=re.IGNORECASE)
    products_match = re.findall(products_pattern, text, flags=re.IGNORECASE)
    seller_match = re.findall(sellers_pattern, text, flags=re.IGNORECASE)
    seller_verbs_match = re.findall(seller_verbs_pattern, text, flags=re.IGNORECASE)
    ad_camp_match = re.findall(ad_camp_pattern, text, flags=re.IGNORECASE)
    negations_match = re.findall(negations_pattern, text, flags=re.IGNORECASE)
    intent_suffixes_match= re.findall(intent_suffixes_pattern, text, flags=re.IGNORECASE)

    B=len(buyer_match)
    S=len(seller_match)
    BV=len(buyer_verbs_match)
    SV=len(seller_verbs_match)
    P=len(products_match)
    Ad=len(ad_camp_match)
    IS=len(intent_suffixes_match)
    PN=len(negations_match)
    
    ws2.cell(row+1,1).value=B
    ws2.cell(row+1,2).value=S
    ws2.cell(row+1,3).value=SV
    ws2.cell(row+1,4).value=BV
    ws2.cell(row+1,5).value=P
    ws2.cell(row+1,6).value=Ad 
    ws2.cell(row+1,7).value=IS
    ws2.cell(row+1,8).value=PN


    # adding label to the training data
    ws2.cell(row+1,9).value=label.upper()

    row+=1
            
    print(f"Processing... {row} ", end='\r')
    wb2.save(r"NLP\dataset_trained.xlsx")
    wb2.close()
# ------------------------- End of Evalute()--------------------------------

if choice==1:
    training_file_txt=input("Waiting for your sentence>>>")
    label=input("Select Labels: B, S, BS or N >>>"  ).strip()
    evaluate(training_file_txt,label)
    print(" 🟢")
    

elif choice==2:
    num=int(input("Number of files >>>  ").strip())
    
    for c in range(num):
        path=input(f"File path #{c+1} >>>").strip()
        label=input("Select Labels:  B, S, BS or N >>>").strip().upper()
        training_files_n_labels.append([path,label])
        
    for f in range(num): # num is number of files that user feeded for training
        
        training_file_txt=training_files_n_labels[f][0]
        label=training_files_n_labels[f][1]
        
        with open(training_file_txt, 'r') as file: # just to find total length for progress bar
            length=len([line for line in file])
        
        
        with open(training_file_txt, 'r') as file:
            progress_bar = tqdm(total=length, unit="iteration") # progress Bar
            for line in file:
                evaluate(line,label)
                progress_bar.update(1)
                
        progress_bar.close()
                
    print("🟢")
        


